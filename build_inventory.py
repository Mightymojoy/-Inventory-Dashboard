"""
ITO 库存看板 · 构建脚本 v0.7
================================
读取  库存-每日.xlsx  的：
  - [ERP库存数据源]      主数据（3320 SKU × 66 列）
  - [供应链提货计划表]    到货时间监控（按日期的提货数量网格）
  - [P1库存渠道分配--专项] 渠道库存（总仓/天猫/达播锁库）

计算库存健康度 / 货值(零售) / 在途 / 到货 等指标，输出 inventory_data.json。

用法：
  python build_inventory.py            # 生成 inventory_data.json 快照
  python build_inventory.py --serve    # 启动本地实时服务（看板用 ?live=1 读取 /api/data）
  python build_inventory.py --port 8765

数据铁律：数据准确 > 省时；无数据则不展示。
"""
import os
import sys
import re
import json
import math
import shutil
import hashlib
import datetime
import subprocess

try:
    import openpyxl
except ImportError:
    sys.exit("需要 openpyxl：请先 pip install openpyxl")

# ---------------- 配置 ----------------
OUT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(OUT_DIR, "库存-每日.xlsx")   # 每日更新表（新格式单 sheet，或含多 sheet 的老格式）
OUT_JSON = os.path.join(OUT_DIR, "inventory_data.json")
ARRIVAL_SYNC_XLSX = os.path.join(OUT_DIR, "入库明细同步.xlsx")   # 到货进度表（按日期命名 sheet，取最新）

# 每日刷新使用当日日期
TODAY = datetime.date.today()

# 常规备货分类（参与库存健康度 / 缺货预警判定）
STOCKED = {"旅行箱", "包袋", "旅行生活"}
# 按需定制分类（可发=0 属正常，不计入缺货 / 预警）
MADE_TO_ORDER = {"定制旅行箱", "定制包袋", "套装"}
# 看板纳入的全部在售分类
ON_SALE = STOCKED | MADE_TO_ORDER

# 健康度阈值（周转天数）
URGENT_DAYS = 7      # 周转 < 7  -> 急需关注
RISK_DAYS = 15       # 周转 < 15 -> 缺货风险（沿用"周转低于15天自动预警"）
OVERSTOCK_DAYS = 90  # 周转 > 90 -> 滞销

LEVEL_ORDER = ["缺货", "急需关注", "缺货风险", "安全", "滞销"]
LEVEL_COLOR = {
    "缺货": "#d9534f",
    "急需关注": "#e8893b",
    "缺货风险": "#e0b53b",
    "安全": "#3a9d6e",
    "滞销": "#7a8794",
    "定制": "#9aa7b0",
}

# ---------------- 工具 ----------------
def num(x):
    """把单元格值安全转成 float，处理 None / '#N/A' / 字符串。"""
    if x is None:
        return 0.0
    if isinstance(x, str):
        s = x.strip()
        if s in ("", "#N/A", "N/A", "NaN", "null"):
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def first_index(header, name):
    """返回列名第一次出现的位置（处理重复列名，颜色/尺寸在数据源出现两次，取 SKU 规格级字段）。"""
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def last_index(header, name):
    """返回列名最后一次出现的位置（兜底用）。"""
    idx = None
    for i, h in enumerate(header):
        if h == name:
            idx = i
    return idx


def pick_attr(row, idx_first, idx_last):
    """合并两个相同字段的有效值：优先取 SKU 级规格字段，空/#N/A 时回退到第二份。"""
    BAD = (None, "", "#N/A", "N/A")
    def _v(i):
        return None if i is None or i >= len(row) else row[i]
    v1 = _v(idx_first)
    if v1 not in BAD:
        return v1
    v2 = _v(idx_last)
    if v2 not in BAD:
        return v2
    return v1 or v2 or ""


def parse_series(name):
    """从货品名称中解析系列，如 'ITO MYCENA BACKPACK系列 ...' -> 'MYCENA BACKPACK'。"""
    if not name:
        return ""
    i = name.find("系列")
    if i <= 0:
        return ""
    pre = name[:i]
    j = pre.rfind("ITO ")
    if j >= 0:
        return pre[j + 4:].strip()
    return pre.strip()


_SIZE_PATTERNS = (
    re.compile(r"(\d+(?:\.\d+)?)\s*英寸"),
    re.compile(r"(\d+(?:\.\d+)?)\s*L(?![A-Za-z])"),   # 注意：不用 \b，中文"包/袋"后无词边界
    re.compile(r"(\d+(?:\.\d+)?)\s*ML"),
    re.compile(r"(\d+(?:\.\d+)?)\s*cm"),
)


def parse_size(name):
    """从商品名称解析尺寸规格：20英寸 / 15L / 600ML / 50cm 等；解析不到返回空串。"""
    if not name:
        return ""
    m = _SIZE_PATTERNS[0].search(name)
    if m:
        return m.group(1) + "英寸"
    m = _SIZE_PATTERNS[1].search(name)
    if m:
        return m.group(1) + "L"
    m = _SIZE_PATTERNS[2].search(name)
    if m:
        return m.group(1) + "ML"
    m = _SIZE_PATTERNS[3].search(name)
    if m:
        return m.group(1) + "cm"
    return ""


def load_matching(wb):
    """读取『匹配表（勿删）』：F列商家编码 -> {颜色(J), 尺寸(K)}。
    匹配表提供正确的 颜色/尺寸（数据源自带的两列语义是 里布材质/颜色名，不可直接用）。"""
    mapping = {}
    try:
        ws = wb["匹配表（勿删）"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or len(r) < 11 or r[5] is None:
                continue
            code = str(r[5]).strip()
            if not code or code in mapping:
                continue  # 保留首次出现
            color = r[9]
            size = r[10]
            if color in (None, "", "#N/A", "N/A"):
                color = ""
            if size in (None, "", "#N/A", "N/A"):
                size = ""
            mapping[code] = {"color": color, "size": size}
    except Exception as e:
        print("[warn] 读取匹配表失败:", e, file=sys.stderr)
    return mapping


# ---------------- 主计算 ----------------
def find_source_sheet(wb):
    """自动识别主数据 sheet：找含『商家编码』表头的 sheet（兼容 新格式 Sheet1 / 老格式 ERP库存数据源）。"""
    for name in wb.sheetnames:
        try:
            row = next(wb[name].iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row and "商家编码" in row:
                return name
        except Exception:
            continue
    return wb.sheetnames[0]


def load_arrival_progress(path=ARRIVAL_SYNC_XLSX):
    """读取『入库明细同步.xlsx』，取**最新日期命名**的 sheet 的交付计划。
    返回 {per_sku, timeline, latest_sheet, latest_date}；无文件/无日期 sheet 返回 None。
    - per_sku: 货品编号 -> {total, by_date:{date:qty}, eta, name}
    - timeline: 日期str -> {qty, skus:[{name,qty}]}（含当天起的交付计划）
    兼容两种表头：第一行直接是日期列（新格式），或第一行'交付节奏'+第二行日期列（老格式）。"""
    if not os.path.exists(path):
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print("[warn] 读取入库明细同步失败:", e, file=sys.stderr)
        return None
    # 找最新日期 sheet（MMDD 命名，当年）
    sheet_dates = {}
    for name in wb.sheetnames:
        if not name or not re.fullmatch(r"\d{4}", name):
            continue
        mm, dd = int(name[:2]), int(name[2:4])
        try:
            d = datetime.date(TODAY.year, mm, dd)
        except ValueError:
            continue
        sheet_dates[d] = name
    if not sheet_dates:
        print("[warn] 入库明细同步中未找到日期命名 sheet", file=sys.stderr)
        return None
    latest = max(sheet_dates)
    ws = wb[sheet_dates[latest]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    # 定位表头行：第一行若非日期交付列，则用第二行
    hdr_idx = 0
    if not any(("交付" in str(v)) for v in (rows[0][2:] if len(rows[0]) > 2 else [])):
        hdr_idx = 1
    if hdr_idx >= len(rows):
        return None
    header = rows[hdr_idx]
    date_cols = {}
    for i, v in enumerate(header):
        s = str(v) if v is not None else ""
        if "交付" in s:
            dstr = s.split("\n")[0].strip()
            if len(dstr) >= 4 and dstr.isdigit():
                try:
                    date_cols[i] = datetime.date(TODAY.year, int(dstr[:2]), int(dstr[2:4]))
                except ValueError:
                    pass
    per_sku = {}
    timeline = {}
    for r in rows[hdr_idx + 1:]:
        if not r or r[0] is None:
            continue
        sku_no = str(r[0]).strip()
        if not sku_no:
            continue
        name = r[1] if len(r) > 1 else ""
        for ci, d in date_cols.items():
            q = num(r[ci]) if ci < len(r) else 0
            if q <= 0:
                continue
            rec = per_sku.setdefault(sku_no, {"total": 0.0, "by_date": {}, "eta": None, "name": name})
            rec["total"] += q
            rec["by_date"][d.isoformat()] = rec["by_date"].get(d.isoformat(), 0) + q
            if rec["eta"] is None or d < rec["eta"]:
                rec["eta"] = d
            g = timeline.setdefault(d.isoformat(), {"qty": 0.0, "skus": []})
            g["qty"] += q
            g["skus"].append({"name": name, "qty": q})
    return {
        "per_sku": per_sku,
        "timeline": timeline,
        "latest_sheet": sheet_dates[latest],
        "latest_date": latest.isoformat(),
    }


def compute(path=XLSX_PATH):
    wb = openpyxl.load_workbook(path, data_only=True)
    has_matching = "匹配表（勿删）" in wb.sheetnames
    matching = load_matching(wb) if has_matching else {}
    has_arrival_plan = "供应链提货计划表" in wb.sheetnames

    # ===== 1. 主数据源 =====
    src_sheet = find_source_sheet(wb)
    ws = wb[src_sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h is not None}

    def gi(name):
        return last_index(header, name)

    f_code = gi("商家编码")
    f_name = gi("货品名称")
    f_short = gi("货品简称")
    f_sku = gi("货品编号")
    f_cat = gi("分类")
    f_brand = gi("品牌")
    f_barcode = gi("条形码")
    f_kf = gi("可发库存")
    f_stock = gi("库存")
    f_purchase = gi("采购在途")
    f_produce = gi("生产在途量")
    f_transfer = gi("调拨在途")
    f_month = gi("月销量")
    f_d7 = gi("7天销量")
    f_d14 = gi("14天销量")
    f_price = gi("零售价")
    f_color = first_index(header, "颜色")
    f_size  = first_index(header, "尺寸")
    f_color2 = last_index(header, "颜色")
    f_size2  = last_index(header, "尺寸")

    skus = []
    for r in rows[1:]:
        cat = r[f_cat] if f_cat is not None else None
        if cat not in ON_SALE:
            continue
        brand = str(r[f_brand]).strip() if f_brand is not None and r[f_brand] is not None else ""
        if brand != "ITO":
            continue   # 只看品牌为 ITO 的商品
        name = r[f_name] if f_name is not None else ""
        kf = num(r[f_kf])
        stock = num(r[f_stock])
        month = num(r[f_month])
        d7 = num(r[f_d7])
        d14 = num(r[f_d14])
        price = num(r[f_price])
        purchase = num(r[f_purchase])
        produce = num(r[f_produce])
        transfer = num(r[f_transfer])
        barcode = r[f_barcode] if f_barcode is not None else None

        # 日均销量：优先 7天/7，回退 14天/14，再回退 月/30
        if d7 > 0:
            daily = d7 / 7.0
        elif d14 > 0:
            daily = d14 / 14.0
        elif month > 0:
            daily = month / 30.0
        else:
            daily = 0.0

        if cat in MADE_TO_ORDER:
            level = "定制"
            turnover = None
            sellout = None
        elif kf <= 0:
            level = "缺货"
            turnover = 0.0
            sellout = None
        elif daily <= 0:
            level = "滞销"
            turnover = float("inf")
            sellout = None
        else:
            turnover = kf / daily
            if turnover < URGENT_DAYS:
                level = "急需关注"
            elif turnover < RISK_DAYS:
                level = "缺货风险"
            elif turnover <= OVERSTOCK_DAYS:
                level = "安全"
            else:
                level = "滞销"
            sellout = TODAY + datetime.timedelta(days=int(math.ceil(turnover)))

        in_transit = purchase + produce + transfer
        value_retail = max(kf, 0.0) * price

        # 颜色/尺寸：有『匹配表（勿删）』优先用匹配表；无匹配表（新格式单 sheet）直接用数据源 颜色/尺寸 列
        code_str = str(r[f_code]).strip() if f_code is not None and r[f_code] is not None else ""
        sku_no = str(r[f_sku]).strip() if f_sku is not None and r[f_sku] is not None else ""
        mt = matching.get(code_str, None)
        mt_has = bool(mt and (mt["color"] or mt["size"]))
        if mt_has:
            color = mt["color"] or pick_attr(r, f_size, f_size2) or ""
            size = mt["size"] or ""
            if size and color and size == color:
                size = ""   # 匹配表尺寸列复制了颜色名（无真实尺寸），视为空
        else:
            color = pick_attr(r, f_color, f_color2) or ""   # 新格式颜色列语义正确；老格式则回退取尺寸列
            if not color or color in ("#N/A", "N/A"):
                color = pick_attr(r, f_size, f_size2) or ""   # 老格式：数据源"尺寸"列实际是颜色名
            size = pick_attr(r, f_size, f_size2) or ""
        if not size:
            size = parse_size(name)   # 匹配表/数据源尺寸缺失时，回退到商品名称解析（如 20L/15L）
        if not size:
            size = ""   # 确实无尺寸概念的产品（水杯/雨伞/套装等）留空展示"-"，不硬凑

        skus.append({
            "code": code_str,
            "sku": sku_no,
            "barcode": str(barcode) if barcode is not None else "",
            "name": name or "",
            "cat": cat or "",
            "series": parse_series(name),
            "color": color,
            "size": size,
            "kf": kf,
            "stock": stock,
            "month": month,
            "d7": d7,
            "price": price,
            "in_transit": in_transit,
            "purchase": purchase,
            "produce": produce,
            "transfer": transfer,
            "daily": round(daily, 2),
            "turnover": (round(turnover, 1) if (turnover is not None and math.isfinite(turnover)) else None),
            "sellout": (sellout.isoformat() if sellout else None),
            "value": round(value_retail, 1),
            "level": level,
            "arrival_qty": 0,      # 后续由提货计划表填充
            "arrival_eta": None,   # 后续填充
        })

    # ===== 1.5 商品 SKU 总览（全量正品目录：在售 + 停采正品）=====
    # 按『匹配表（勿删）』口径设计字段：商品编码/货品编号/货品名称/货品简称/零售价 + 分类/系列
    NON_GOODS = {"服务", "会员增值服务", "无", "NA", "RAW PROJECT"}
    CATALOG_CATS = ON_SALE | {"停采旅行箱", "停采包袋"}
    ACCESSORY_CATS = {"旅行箱配件", "停采旅行箱配件", "赠品", "停采赠品"}
    catalog = []
    accessories = []
    for r in rows[1:]:
        cat = r[f_cat] if f_cat is not None else None
        if cat is None:
            continue
        brand = str(r[f_brand]).strip() if f_brand is not None and r[f_brand] is not None else ""
        if brand != "ITO":
            continue   # 只看品牌为 ITO 的商品
        name = r[f_name] if f_name is not None else ""
        if not name or cat in NON_GOODS:
            continue
        code = str(r[f_code]).strip() if r[f_code] is not None else ""
        sku_no = str(r[f_sku]).strip() if f_sku is not None and r[f_sku] is not None else ""
        short = str(r[f_short]).strip() if f_short is not None and r[f_short] is not None else ""
        price = num(r[f_price])
        item = {
            "code": code,
            "sku": sku_no,
            "name": name,
            "short": short or "",
            "cat": cat or "",
            "series": parse_series(name),
            "price": price,
        }
        if cat in CATALOG_CATS:
            item["status"] = "在售" if cat in ON_SALE else ("定制" if cat in MADE_TO_ORDER else "停采")
            catalog.append(item)
        elif cat in ACCESSORY_CATS:
            item["kf"] = num(r[f_kf]) if f_kf is not None else 0
            item["stock"] = num(r[f_stock]) if f_stock is not None else 0
            item["status"] = "停采" if str(cat).startswith("停采") else ""
            accessories.append(item)

    # ===== 2. 到货进度 =====
    # 优先级：① 入库明细同步.xlsx（取最新日期 sheet，按货品编号关联）② 供应链提货计划表（按条形码关联）
    global_by_date = {}        # date_str -> {qty, skus:[]}
    future_week_total = 0.0
    arrival_ready = False
    arrival_source = ""

    # ① 入库明细同步（优先）
    ap = load_arrival_progress()
    if ap is not None and ap["per_sku"]:
        arrival_ready = True
        arrival_source = "入库明细同步·" + ap["latest_sheet"] + "（" + ap["latest_date"] + " 更新）"
        print(f"[info] 到货进度: {arrival_source}", file=sys.stderr)
        for s in skus:
            rec = ap["per_sku"].get(s.get("sku", ""))
            if rec:
                s["arrival_qty"] = round(rec["total"], 1)
                s["arrival_eta"] = rec["eta"].isoformat() if rec["eta"] else None
        for d, g in sorted(ap["timeline"].items()):
            global_by_date[d] = g
            days = (datetime.date.fromisoformat(d) - TODAY).days
            if 0 < days <= 7:
                future_week_total += g["qty"]
    # ② 供应链提货计划表（回退）
    elif has_arrival_plan:
        try:
            wa = wb["供应链提货计划表"]
            war = list(wa.iter_rows(values_only=True))
            # 表头在 war[2]：A=条码(0) B=名称(1) C=未来一周提货计划(2) D..=日期
            header_row = war[2]
            date_cols = {}
            for i, v in enumerate(header_row):
                if isinstance(v, (datetime.datetime, datetime.date)):
                    date_cols[i] = v.date() if isinstance(v, datetime.datetime) else v
            arrivals_by_barcode = {}
            # 数据从 war[3] 开始
            for r in war[3:]:
                if not r or r[0] is None:
                    continue
                bc = str(r[0])
                for ci, d in date_cols.items():
                    q = num(r[ci]) if ci < len(r) else 0
                    if q > 0 and d > TODAY:
                        rec = arrivals_by_barcode.setdefault(bc, {"total": 0.0, "by_date": {}, "eta": None})
                        rec["total"] += q
                        rec["by_date"][d.isoformat()] = rec["by_date"].get(d.isoformat(), 0) + q
                        if rec["eta"] is None or d < rec["eta"]:
                            rec["eta"] = d
                        g = global_by_date.setdefault(d.isoformat(), {"qty": 0.0, "skus": []})
                        g["qty"] += q
                        g["skus"].append({"name": r[1] if len(r) > 1 else "", "qty": q})
                        if (d - TODAY).days <= 7:
                            future_week_total += q
            if arrivals_by_barcode:
                arrival_ready = True
                arrival_source = "供应链提货计划表"
                for s in skus:
                    rec = arrivals_by_barcode.get(s["barcode"])
                    if rec:
                        s["arrival_qty"] = round(rec["total"], 1)
                        s["arrival_eta"] = rec["eta"].isoformat() if rec["eta"] else None
        except Exception as e:
            print("[warn] 读取提货计划表失败:", e, file=sys.stderr)
    else:
        print("[info] 未检测到到货进度数据（入库明细同步 / 供应链提货计划表），到货日历不可用", file=sys.stderr)

    # ===== 3. 汇总指标 =====
    kpi = {
        "on_sale_sku": len(skus),
        "custom_sku": sum(1 for s in skus if s["level"] == "定制"),
        "total_kf": round(sum(s["kf"] for s in skus), 1),
        "total_value": round(sum(s["value"] for s in skus), 1),
        "out_of_stock": sum(1 for s in skus if s["level"] == "缺货"),
        "warning": sum(1 for s in skus if s["level"] in ("急需关注", "缺货风险")),
        "in_transit": round(sum(s["in_transit"] for s in skus), 1),
        "arrival_week": round(future_week_total, 1),
    }

    health = {lv: {"count": 0, "value": 0.0} for lv in LEVEL_ORDER}
    for s in skus:
        if s["level"] in health:
            health[s["level"]]["count"] += 1
            health[s["level"]]["value"] += s["value"]

    categories = {}
    for s in skus:
        c = categories.setdefault(s["cat"], {"sku": 0, "value": 0.0, "oos": 0, "warn": 0})
        c["sku"] += 1
        c["value"] += s["value"]
        if s["level"] == "缺货":
            c["oos"] += 1
        if s["level"] in ("急需关注", "缺货风险"):
            c["warn"] += 1

    # 到货按日期排序
    arrival_timeline = [
        {"date": d, "qty": round(g["qty"], 1),
         "skus": sorted(g["skus"], key=lambda x: -x["qty"])[:8]}
        for d, g in sorted(global_by_date.items())
    ]

    result = {
        "version": "0.11",
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_date": TODAY.isoformat(),
        "has_arrival_plan": arrival_ready,
        "arrival_source": arrival_source,
        "src_sheet": src_sheet,
        "config": {
            "on_sale": sorted(ON_SALE),
            "stocked": sorted(STOCKED),
            "made_to_order": sorted(MADE_TO_ORDER),
            "urgent_days": URGENT_DAYS,
            "risk_days": RISK_DAYS,
            "overstock_days": OVERSTOCK_DAYS,
            "level_order": LEVEL_ORDER,
            "level_color": LEVEL_COLOR,
        },
        "kpi": kpi,
        "health": health,
        "categories": categories,
        "arrival_timeline": arrival_timeline,
        "catalog": catalog,
        "accessories": accessories,
        "skus": skus,
    }
    return result


# ---------------- 输出 / 服务 ----------------
def write_json(result, path=OUT_JSON):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[ok] 已生成 {path}  (在售SKU {result['kpi']['on_sale_sku']}, 货值 ￥{result['kpi']['total_value']:,.0f})")


def bundle(result, password=""):
    """生成自包含成品：把数据内嵌进 HTML，双击即可直接打开（无需服务器/无需额外操作）。
    传入 password 时注入登录门（SHA-256 哈希，不落明文）。"""
    tpl = os.path.join(OUT_DIR, "inventory_dashboard.html")
    with open(tpl, encoding="utf-8") as f:
        html = f.read()
    payload = json.dumps(result, ensure_ascii=False)
    script = "<script>window.__DATA__ = " + payload + ";</script>"
    if "<!--__EMBED_DATA__-->" in html:
        html = html.replace("<!--__EMBED_DATA__-->", script)
    else:  # 兜底：插到 body 开头
        html = html.replace("<body>", "<body>\n" + script, 1)
    # 登录门密码哈希（不落明文）
    pwd_hash = hashlib.sha256(password.encode("utf-8")).hexdigest() if password else ""
    pwd_script = ("<script>window.__PWD_HASH__ = '" + pwd_hash + "';</script>") if pwd_hash else "<script>window.__PWD_HASH__ = null;</script>"
    html = html.replace("<!--__PWD_HASH__-->", pwd_script)
    out = os.path.join(OUT_DIR, "ITO库存看板.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    # 注入更新时间轴（并入 build 流程：无论手动跑 / 双击 bat / 定时任务都不会再丢时间轴）
    # 2026-08-19 事故根因：手动跑 build 绕过了 bat 的 3.5/4 步，重新生成的 HTML 覆盖了时间轴
    tl_script = os.path.join(r"C:\Users\QwQ\WorkBuddy\2026-08-12-13-26-00\update_timeline", "update_timeline.py")
    try:
        r = subprocess.run(
            [sys.executable, tl_script, out, "--status", "ok", "--note", "库存看板数据更新"],
            capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30)
        if r.returncode == 0:
            print("[ok] 已注入更新时间轴")
        else:
            print(f"[warn] 更新时间轴注入失败: {r.stderr[:200]}", file=sys.stderr)
    except Exception as e:
        print(f"[warn] 更新时间轴注入异常（不影响看板）: {e}", file=sys.stderr)
    # 同步到 deploy/（Vercel 部署源，Python 内完成，不依赖 cmd copy）
    # 注意：线上 index.html 通过 fetch('inventory_data.json') 取数，
    # 必须连同数据文件一起同步，否则线上取数 404 → 数据永远不更新
    deploy_dir = os.path.join(OUT_DIR, "deploy")
    if os.path.isdir(deploy_dir):
        try:
            shutil.copy(out, os.path.join(deploy_dir, "index.html"))
            print("[ok] 已同步 deploy/index.html（Vercel 部署源）")
        except Exception as e:
            print(f"[warn] deploy/index.html 同步失败（不影响本地看板）: {e}", file=sys.stderr)
        try:
            shutil.copy(OUT_JSON, os.path.join(deploy_dir, "inventory_data.json"))
            print("[ok] 已同步 deploy/inventory_data.json（线上取数依赖此文件）")
        except Exception as e:
            print(f"[warn] deploy/inventory_data.json 同步失败（线上将取数 404）: {e}", file=sys.stderr)
    if password:
        print(f"[ok] 已生成自包含成品 {out}  （含密码登录门）")
    else:
        print(f"[ok] 已生成自包含成品 {out}  （直接双击打开即可）")


def get_password():
    """获取登录密码：优先 --password 参数，否则读同目录 password.txt（.gitignore 排除，不入库）。"""
    if "--password" in sys.argv:
        try:
            pw = sys.argv[sys.argv.index("--password") + 1].strip()
            if pw:
                return pw
        except IndexError:
            pass
    pw_file = os.path.join(OUT_DIR, "password.txt")
    if os.path.exists(pw_file):
        pw = open(pw_file, encoding="utf-8").read().strip()
        if pw:
            return pw
    return ""


def serve(port=8765):
    """本地实时服务：GET /api/data 重新读 xlsx 计算；GET / 返回看板。"""
    import http.server
    import functools

    html_path = os.path.join(OUT_DIR, "inventory_dashboard.html")
    script_dir = OUT_DIR

    class Handler(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            if self.path.startswith("/api/data"):
                try:
                    data = compute()
                    body = json.dumps(data, ensure_ascii=False).encode("utf-8")
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Access-Control-Allow-Origin", "*")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                except Exception as e:
                    self.send_response(500)
                    self.end_headers()
                    self.wfile.write(str(e).encode("utf-8"))
            elif self.path in ("/", "/index.html"):
                try:
                    with open(html_path, "rb") as f:
                        body = f.read()
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html; charset=utf-8")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                except FileNotFoundError:
                    self.send_response(404)
                    self.end_headers()
            else:
                self.send_response(404)
                self.end_headers()

        def log_message(self, *args):
            pass

    handler = functools.partial(Handler)
    httpd = http.server.ThreadingHTTPServer(("127.0.0.1", port), handler)
    print(f"[serve] 实时服务已启动: http://127.0.0.1:{port}/  (Ctrl+C 停止)")
    print(f"[serve] 看板访问: http://127.0.0.1:{port}/inventory_dashboard.html?live=1")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n[serve] 已停止")


if __name__ == "__main__":
    if "--serve" in sys.argv:
        p = 8765
        if "--port" in sys.argv:
            try:
                p = int(sys.argv[sys.argv.index("--port") + 1])
            except (ValueError, IndexError):
                pass
        serve(p)
    else:
        xlsx = XLSX_PATH
        if "--file" in sys.argv:   # 支持自定义数据源：python build_inventory.py --file xxx.xlsx
            try:
                xlsx = sys.argv[sys.argv.index("--file") + 1]
            except IndexError:
                pass
        if not os.path.exists(xlsx):
            sys.exit(f"[错误] 找不到数据源文件: {xlsx}\n请确认每日更新的「库存-每日.xlsx」已在 {OUT_DIR} 目录下。")
        res = compute(xlsx)
        write_json(res)
        bundle(res, password=get_password())
